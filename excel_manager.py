"""
Excel dosya yönetimi - Mükerrer kontrol ve güncelleme
"""
import os
import hashlib
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from logger import setup_logger

logger = setup_logger(__name__)


class ExcelManager:
    """Excel dosya yönetimi ve mükerrer kontrol sınıfı"""
    
    def __init__(self, excel_path=None):
        """
        Excel Manager başlatıcı
        
        Args:
            excel_path: Varolan Excel dosya yolu (opsiyonel)
        """
        self.excel_path = excel_path
        self.existing_hashes = set()
        self.existing_data = []
        
        if excel_path and os.path.exists(excel_path):
            self._load_existing_data()
            logger.info(f"Varolan Excel yüklendi: {excel_path} ({len(self.existing_data)} kayıt)")
        else:
            logger.info("Yeni Excel dosyası oluşturulacak")
    
    def _load_existing_data(self):
        """Varolan Excel dosyasını yükler"""
        try:
            # Excel'i oku
            df = pd.read_excel(self.excel_path)
            self.existing_data = df.to_dict('records')
            
            # Hash'leri hesapla (mükerrer kontrol için)
            for row in self.existing_data:
                file_hash = self._calculate_hash(row)
                self.existing_hashes.add(file_hash)
            
            logger.info(f"Varolan {len(self.existing_data)} kayıt yüklendi")
            
        except Exception as e:
            logger.error(f"Excel yükleme hatası: {str(e)}")
            self.existing_data = []
            self.existing_hashes = set()
    
    def _calculate_hash(self, data_dict):
        """
        Fiş için unique hash hesaplar (mükerrer kontrol)
        
        Args:
            data_dict: Fiş verisi
            
        Returns:
            str: Hash değeri
        """
        try:
            # Kritik alanları birleştir
            key_fields = [
                str(data_dict.get('Dosya Adı', '')),
                str(data_dict.get('Firma', '')),
                str(data_dict.get('Tutar', '')),
                str(data_dict.get('Tarih', ''))
            ]
            
            combined = '|'.join(key_fields)
            return hashlib.md5(combined.encode()).hexdigest()
            
        except Exception as e:
            logger.error(f"Hash hesaplama hatası: {str(e)}")
            return ''
    
    def is_duplicate(self, data_dict):
        """
        Fişin daha önce eklenip eklenmediğini kontrol eder
        
        Args:
            data_dict: Kontrol edilecek fiş verisi
            
        Returns:
            bool: True ise mükerrer, False ise yeni
        """
        file_hash = self._calculate_hash(data_dict)
        is_dup = file_hash in self.existing_hashes
        
        if is_dup:
            logger.warning(f"Mükerrer kayıt tespit edildi: {data_dict.get('Dosya Adı')}")
        
        return is_dup
    
    def add_records(self, new_data_list):
        """
        Yeni kayıtları ekler (mükerrer kontrolü ile)
        
        Args:
            new_data_list: Yeni fiş verileri listesi
            
        Returns:
            tuple: (eklenen_sayı, mükerrer_sayı)
        """
        added_count = 0
        duplicate_count = 0
        
        for data in new_data_list:
            if not self.is_duplicate(data):
                self.existing_data.append(data)
                file_hash = self._calculate_hash(data)
                self.existing_hashes.add(file_hash)
                added_count += 1
            else:
                duplicate_count += 1
        
        logger.info(f"Eklenen: {added_count}, Mükerrer: {duplicate_count}")
        return added_count, duplicate_count
    
    def save_to_excel(self, output_path=None, styled=True):
        """
        Verileri Excel'e kaydeder
        
        Args:
            output_path: Çıktı dosya yolu (None ise mevcut dosya)
            styled: Stil uygula (renkler, formatlar)
            
        Returns:
            str: Oluşturulan dosya yolu
        """
        try:
            if output_path is None:
                output_path = self.excel_path or f"fis_raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            logger.info(f"Excel kaydediliyor: {output_path}")
            
            # DataFrame oluştur
            df = pd.DataFrame(self.existing_data)
            
            if styled:
                # Stil uygulayarak kaydet
                self._save_styled_excel(df, output_path)
            else:
                # Basit kayıt
                df.to_excel(output_path, index=False)
            
            logger.info(f"Excel kaydedildi: {output_path} ({len(self.existing_data)} kayıt)")
            return output_path
            
        except Exception as e:
            logger.error(f"Excel kaydetme hatası: {str(e)}")
            return None
    
    def _save_styled_excel(self, df, output_path):
        """
        Stil uygulanmış Excel kaydeder
        
        Args:
            df: pandas DataFrame
            output_path: Çıktı dosya yolu
        """
        try:
            # Excel'e yaz
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Fişler', index=False)
                
                # Workbook ve sheet'i al
                workbook = writer.book
                worksheet = writer.sheets['Fişler']
                
                # Başlık stili
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Border
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Başlıkları stillendir
                for col_num, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    
                    # Sütun genişliği
                    column_letter = cell.column_letter
                    worksheet.column_dimensions[column_letter].width = 20
                
                # Veri hücrelerini stillendir
                for row_num in range(2, len(df) + 2):
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="center")
                        
                        # AI analizi olan satırları vurgula
                        if col_num == df.columns.get_loc('Durum') + 1:
                            if 'AI Analiz' in str(cell.value):
                                cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                                cell.font = Font(color="155724", bold=True)
                
                # Özet sayfası ekle
                self._add_summary_sheet(workbook, df)
                
            logger.info("Stil uygulaması tamamlandı")
            
        except Exception as e:
            logger.error(f"Stil uygulama hatası: {str(e)}")
            # Hata olursa basit kaydet
            df.to_excel(output_path, index=False)
    
    def _add_summary_sheet(self, workbook, df):
        """
        Özet sayfası ekler
        
        Args:
            workbook: openpyxl Workbook
            df: pandas DataFrame
        """
        try:
            # Özet sheet oluştur
            summary_sheet = workbook.create_sheet('Özet')
            
            # Başlık
            summary_sheet['A1'] = 'FIŞ RAPORU ÖZETİ'
            summary_sheet['A1'].font = Font(bold=True, size=16, color="366092")
            summary_sheet.merge_cells('A1:B1')
            
            # İstatistikler
            stats = [
                ('', ''),
                ('Toplam Fiş Sayısı', len(df)),
                ('Başarılı İşlem', len(df[df['Durum'].str.contains('Başarılı', na=False)])),
                ('AI ile Analiz', len(df[df['Durum'].str.contains('AI Analiz', na=False)])),
                ('', ''),
                ('Toplam Tutar', f"{df['Tutar'].sum():.2f} TL"),
                ('Ortalama Tutar', f"{df['Tutar'].mean():.2f} TL"),
                ('En Yüksek Tutar', f"{df['Tutar'].max():.2f} TL"),
                ('En Düşük Tutar', f"{df['Tutar'].min():.2f} TL"),
                ('', ''),
            ]
            
            for idx, (label, value) in enumerate(stats, 3):
                summary_sheet.cell(row=idx, column=1).value = label
                summary_sheet.cell(row=idx, column=2).value = value
                
                if label:  # Boş satırları atla
                    summary_sheet.cell(row=idx, column=1).font = Font(bold=True)
            
            # Tür dağılımı
            row = len(stats) + 4
            summary_sheet.cell(row=row, column=1).value = 'TÜR DAĞILIMI'
            summary_sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 1
            
            type_counts = df['Tür'].value_counts()
            for receipt_type, count in type_counts.items():
                row += 1
                summary_sheet.cell(row=row, column=1).value = receipt_type
                summary_sheet.cell(row=row, column=2).value = count
            
            # Sütun genişliği
            summary_sheet.column_dimensions['A'].width = 25
            summary_sheet.column_dimensions['B'].width = 20
            
            logger.info("Özet sayfası eklendi")
            
        except Exception as e:
            logger.error(f"Özet sayfası ekleme hatası: {str(e)}")
    
    def get_statistics(self):
        """
        İstatistikleri döndürür
        
        Returns:
            dict: İstatistikler
        """
        if not self.existing_data:
            return {}
        
        df = pd.DataFrame(self.existing_data)
        
        return {
            'total_count': len(df),
            'total_amount': df['Tutar'].sum(),
            'avg_amount': df['Tutar'].mean(),
            'ai_analyzed': len(df[df['Durum'].str.contains('AI Analiz', na=False)]),
            'unique_companies': df['Firma'].nunique(),
            'type_distribution': df['Tür'].value_counts().to_dict()
        }

